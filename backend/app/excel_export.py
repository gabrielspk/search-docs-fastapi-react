import openpyxl
import os
from tkinter import messagebox
from utils.utilitarios import incrementar_nome_arquivo

def criar_relatorio(documentos_encontrados, documentos_nao_encontrados, qtde_arquivos_pesquisados, nome_arquivo, diretorio_atual): #função responsável por criar relatório dos arquivos pesquisados
    workbook = openpyxl.Workbook() #declarando o workbook
    sheet = workbook.active #ativando o workbook do openpyxl

    #inicializando as colunas principais
    sheet['A1'] = 'Documentos em arquivos' 
    sheet['B1'] = 'Nomenclatura arquivo'
    sheet['C1'] = 'Não encontrado'

    #declarando as linhas dos documentos encontrados e não encontrados
    linha_docs_encontrados = 1
    linha_docs_nao_encontrados = 1

    #iniciando laço para percorrer os documentos encontrados
    for documento, arquivo in documentos_encontrados:
        linha_docs_encontrados += 1
        sheet.cell(row=linha_docs_encontrados, column=1, value=documento)
        sheet.cell(row=linha_docs_encontrados, column=2, value=arquivo)

    #laço que percorre os documentos não encontrados
    for documento in documentos_nao_encontrados:
        linha_docs_nao_encontrados += 1 #somando na variável para virar "C2" e sucessivamente.
        sheet.cell(row=linha_docs_nao_encontrados, column=3).value = documento #aplicando o documento não encontrado na linha atual e coluna 3

    #mudando a aplicação para o diretório escolhido pelo usuário
    os.chdir(diretorio_atual)
    nome_arquivo = incrementar_nome_arquivo(nome_arquivo) #chamando a função que incrementa na nomenclatura do arquivo
    
    caminho_relatorio = os.path.join(diretorio_atual, nome_arquivo) #juntando o diretório com o nome do arquivo para salvar o mesmo 

    #tentando salvar o relatório, se estiver aberto será retorno uma except
    try:
        workbook.save(caminho_relatorio)
        return f"Relatório gerado com sucesso: {nome_arquivo}"
    except PermissionError:
        return "Não foi possível gerar o relatório, pois está aberto por outro programa."